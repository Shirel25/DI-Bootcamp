# Import the openpyxl module
from openpyxl import load_workbook
# Avoid header warnings 
import warnings
warnings.simplefilter("ignore")


# Open the workbook containing the list of plants
plant_book = load_workbook(filename=r"C:\Users\shire\OneDrive\Fac\DevelopersInstitute\week6\day2\ExercicesXP\ex2\Plants.xlsx")

# Get a reference to the first sheet (called Sheet1)
plant_sheet = plant_book["Sheet1"]

# Start with cell A1 
this_plant = plant_sheet.cell(1,1)

# Initializing a counter to avoid infinite loops
int_check = 0
while True:
     # avoid infinite loops if bug
    int_check += 1
    if int_check > 100:
        print("Integer check triggered")
        break
    # go to next plant (bottom line)
    this_plant = this_plant.offset(1,0)

    # if no such plant, stop
    if this_plant.value == None:
        print("\nThe above plants are not in stock")
        break

    # if this plant is not in stock, say so
    stock_cell = this_plant.offset(0,7)
    if stock_cell.value == "No":
        print(this_plant.value)

# close down the file
plant_book.close()
